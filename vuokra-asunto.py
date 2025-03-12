
import datetime
from bs4 import BeautifulSoup
import requests
import openpyxl
import re
import random
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager

url = "https://asunnot.oikotie.fi/vuokra-asunnot?pagination=1&locations=%5B%5B6610803,4,%22L%C3%A4nsi-Helsinki,%20Helsinki%22%5D%5D&price%5Bmax%5D=1390&size%5Bmin%5D=48&roomCount%5B%5D=2&buildingType%5B%5D=1&buildingType%5B%5D=256&cardType=101"
currentdate = datetime.datetime.now()
listings = []

# List of User-Agent strings to choose from
user_agents = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/94.0.4606.81 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:92.0) Gecko/20100101 Firefox/92.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Version/14.0.3 Safari/537.36",
    "Mozilla/5.0 (Linux; Android 10; SM-G970F) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.120 Mobile Safari/537.36",
    "Mozilla/5.0 (iPhone; CPU iPhone OS 14_0 like Mac OS X) AppleWebKit/537.36 (KHTML, like Gecko) Version/14.0 Mobile/15E148 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36 Edg/91.0.864.59",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.212 Safari/537.36 Edg/90.0.818.56"
]

def get_house_info(url):

    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')
    house_info = {}

    for item in soup.find_all('div', class_='details-grid__item'):
        # Find the <dt> and <dd> within each item
        label = item.find('dt', class_='details-grid__item-title')
        value = item.find('dd', class_='details-grid__item-value')

        if label and value:
            house_info[label.text.strip()] = value.text.strip()

    # Now assign each value to a separate variable
    house_info['vuokra_kk'] = house_info.get('Vuokra/kk', 'Not available')
    house_info['asuinpinta_ala'] = house_info.get('Asuinpinta-ala', 'Not available')
    house_info['huoneita'] = house_info.get('Huoneita', 'Not available')
    house_info['kerros'] = house_info.get('Kerros', 'Not available')
    house_info['rakennusvuosi'] = house_info.get('Rakennusvuosi', 'Not available')
    house_info['rakennuksen_tyyppi'] = house_info.get('Rakennuksen tyyppi', 'Not available')
    house_info['kaupunginosa'] = house_info.get('Kaupunginosa', 'Not available')
    house_info['kaupunki'] = house_info.get('Kaupunki', 'Not available')

    return house_info

def send_telegram_message(message):
    api_token = "-insert api token here--"
    chat_id = "chat id here ----"

    response = requests.post(f'https://api.telegram.org/bot{api_token}/sendMessage',
                             data={'chat_id': chat_id, 'text': message})

    if response.status_code == 200:
        print("Hyvä ilmoitus löydetty ja viesti lähetetty.")
    else:
        print(
            f"Hyvä ilmoitus löydetty, mutta viesti ei mennyt perille. ErrorCode: {response.status_code} - {response.content}")


def main():

    #Selenium setup
    options = Options()
    random_user_agent = random.choice(user_agents)
    options.add_argument("--headless")  # Run in headless mode (no UI)
    options.add_argument(f"user-agent={random_user_agent}")  # Set the randomized user-agent
    options.add_argument("--disable-gpu")  # Disable GPU (good for headless mode)
    options.add_argument("--window-size=1920,1080")  # Set window size for the headless browser

    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    driver.get(url)
    time.sleep(3)


    surl = driver.find_elements(By.XPATH, "//a[contains(@class, 'ot-card') and contains(@class, 'link--muted')]")
    #listings = ["https://asunnot.oikotie.fi/vuokra-asunnot/helsinki/22500541"]

    for y in surl:
        listings.append(y.get_attribute("href"))
    print(listings)
    driver.quit()

    excel_file_path = "C:\\tests\\listings.xlsx"
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active

    for apartment in listings:
        print(apartment)
        if apartment in [cell.value for cell in sheet['B']]:
            print("Löytyy jo, skipataan seuraavaan")
            continue

        house_info = get_house_info(apartment)
        vuokra = re.sub(r'\D', '', house_info['vuokra_kk'])
        asuinpinta_ala = re.sub(r'\D', '', house_info['asuinpinta_ala'])

        neliohinta = round(int(vuokra)/int(asuinpinta_ala))

        row = sheet.max_row + 1
        sheet.cell(row=row, column=1, value=currentdate)
        sheet.cell(row=row, column=2, value=apartment)
        sheet.cell(row=row, column=3, value=house_info['vuokra_kk'])
        sheet.cell(row=row, column=4, value=house_info['asuinpinta_ala'])
        sheet.cell(row=row, column=5, value=house_info['asuinpinta_ala'])
        sheet.cell(row=row, column=6, value=house_info['kerros'])
        sheet.cell(row=row, column=7, value=house_info['rakennusvuosi'])
        sheet.cell(row=row, column=8, value=house_info['kaupunginosa'])

        message = (f"Uusi ilmoitus!\nHinta: {house_info['vuokra_kk']}\n"
                   f"Neliöhinta: {neliohinta} €\n"
                   f"Alue: {house_info['kaupunginosa']}\n"
                   f"Rakennusvuosi: {house_info['rakennusvuosi']}\n"
                   f"URL: {apartment}")

        send_telegram_message(message)

    workbook.save(excel_file_path)
    workbook.close()


if __name__ == "__main__":
    main()
