# Asuntoetsijä

# Allows you to send HTTP requests using Python
import requests

# Verkkotiedon löytäminen HTML & XML tiedostoista
from bs4 import BeautifulSoup

# Excel tiedoston muokkaaminen
import openpyxl

# Tarvitaan, jotta saadaan muutettua hinta ja pinta-ala int muotoon
import re



# Asuntosivuston (etuovi) URL ja siihen beautifulsoupin vastaukset
url1 = 'https://www.etuovi.com/myytavat-asunnot?haku=M2061581391'
vastaus = requests.get(url1)
html = vastaus.content
soup = BeautifulSoup(html, 'html.parser')

# Luodaan lista johon voidaan listata hakutulokset listatuista hakutuloksista
talot = []

# Etsitään kaikki sivulla näkyvät ilmoitukset ja niiden URL. Lisätään hakutulokset listaan
kohteet = soup.find_all(class_="mui-style-wp78j0 e1re311t2")
for kohde in kohteet:
    if kohde is not None:
        kohdeURL = kohde["href"]
        #print(kohdeURL)
        talot.append(kohdeURL)

# Exceltiedosto
sijainti = "C:\\Users\\----------"
exceltiedosto = openpyxl.load_workbook(sijainti)
sheet = exceltiedosto.active

#print(talot)

# Haetaan ilmoituksen kaikki tiedot ja lisätään ne Exceltiedostoon
for x in talot:
        if x in [cell.value for cell in sheet["B"]]:
                print("löytyy jo, skipataan") 
                continue
       
       # Lisätään Exceltiedostoon URL
        row = sheet.max_row + 1
        sheet.cell(row=row, column=2, value=x)
        print(f"lisätty ilmoitus: {x}")

        taloURL = ("https://www.etuovi.com/" + x)

        # Haetaan ilmoituksen URL josta löydetään ilmoituksen tiedot
        vastausINNER = requests.get(taloURL)
        htmlINNER = vastausINNER.content
        soupINNER = BeautifulSoup(htmlINNER, 'html.parser')

        # Haetaan hinta
        hintaluokka = soupINNER.find_all(class_="MuiGrid-root MuiGrid-item MuiGrid-grid-xs-4 MuiGrid-grid-md-5 mui-style-1ymh2wi")
        #print(hintaluokka)
        for xhinta in hintaluokka:
                hinta1 = xhinta.find("h3").text
                #print(hinta1)
                sheet.cell(row=row, column=3, value=hinta1)
                
        
        # Haetaan rakennusvuosi
        rakennusvuosi = soupINNER.find_all(class_="MuiGrid-root MuiGrid-item MuiGrid-grid-xs-4 MuiGrid-grid-md-3 mui-style-1niyv08")
        #print(rakennusvuosi)
        for xvuosi in rakennusvuosi:
                rakennusvuosi1 = xvuosi.find("h3").text
                #print(rakennusvuosi1) 
                sheet.cell(row=row, column=4, value=rakennusvuosi1)

        # Haetaan osoite
        osoiteluokka = soupINNER.find_all(class_="MuiGrid-root MuiGrid-item MuiGrid-grid-xs-12 MuiGrid-grid-md-6 oqCVsVj mui-style-1bi94kt")
        #print(osoiteluokka)
        for xosoite in osoiteluokka:
                hinta3 = xosoite.find("h1").text
                #print(hinta3)
                sheet.cell(row=row, column=5, value=hinta3)

        # Haetaan asunnon pinta-ala neliömetreinä
        asunnonkoko = soupINNER.find_all(class_="MuiGrid-root MuiGrid-item MuiGrid-grid-xs-4 MuiGrid-grid-md-4 mui-style-j3iqgs")
        #print(asunnonkoko)
        for xkoko in asunnonkoko:
                hinta4 = xkoko.find("span").text
                #print(hinta4)
                sheet.cell(row=row, column=6, value=hinta4)

        # Haetaan huonemäärä
        huonemäärä = soupINNER.find("h2", {"MuiGrid-root MuiGrid-item MuiGrid-grid-xs-12 MuiGrid-grid-md-6 oqCVsVj mui-style-1bi94kt"})
        huonemäärä1 = huonemäärä.text.strip()
        #print(huonemäärä1)
        for xhuone in huonemäärä:
                sheet.cell(row=row, column=7, value=huonemäärä1)

        # Lasketaan neliöhinta
        # Muutetaan hinta ja pinta-ala int muotoon
        hinta11 = int(''.join(filter(str.isdigit, hinta1)))
        hinta44 = int(re.search(r'\d+', hinta4).group())
        neliöhinta = hinta11 / hinta44
        neliöhinta1 = round(neliöhinta)
        sheet.cell(row=row, column=8, value=neliöhinta1)

        rakennusvuosi11 = int(''.join(filter(str.isdigit, rakennusvuosi1)))

        # Lähetetään viesti Telegrammin kautta, mikäli ilmoitus on tarpeeksi hyvä
        if hinta11 <= 80000 and hinta44 <= 2000 and rakennusvuosi11 >= 1950:
                # Tiedot, jotka tarvitaan viestin lähettämiseen
                api_token = "---apitoken-----"
                chat_id = "---chatid"

                message = ("Uusi ilmoitus! \n Hinta: " + hinta1 + "\n Rakennusvuosi: " + rakennusvuosi1 + "\n Neliöhinta: " + str(neliöhinta1) + "€ \n URL: " + taloURL)

                # Lähetetään viesti
                response = requests.post(f'https://api.telegram.org/bot{api_token}/sendMessage',
                                        data={'chat_id': chat_id, 'text': message})
                
                # Tarkistus
                if response.status_code == 200:
                        print("Hyvä ilmoitus löydetty ja viesti lähetetty.")
                else:
                        print(f"Hyvä ilmoitus löydetty, mutta viesti ei mennyt perille. ErrorCode: {response.status_code} - {response.content}")

exceltiedosto.save(sijainti)



