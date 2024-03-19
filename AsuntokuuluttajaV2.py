from datetime import datetime
from bs4 import BeautifulSoup
import requests
import openpyxl
import re


def get_house_info(url):

    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')
    house_info = {}

    price1 = soup.find(class_='MuiGrid-root MuiGrid-item MuiGrid-grid-xs-4 MuiGrid-grid-md-5 mui-style-1ymh2wi')
    year1 = soup.find(class_='MuiGrid-root MuiGrid-item MuiGrid-grid-xs-4 MuiGrid-grid-md-3 mui-style-1niyv08')
    address1 = soup.find(class_='MuiGrid-root MuiGrid-item MuiGrid-grid-xs-12 MuiGrid-grid-md-6 YPV62q_ mui-style-1bi94kt')
    area1 = soup.find(class_='MuiGrid-root MuiGrid-item MuiGrid-grid-xs-4 MuiGrid-grid-md-4 mui-style-j3iqgs')
    rooms1 = soup.find(class_='MuiGrid-root MuiGrid-item MuiGrid-grid-xs-12 MuiGrid-grid-md-6 YPV62q_ mui-style-1bi94kt')
    house_info['price'] = price1.find('h3').text.strip()
    house_info['year_built'] = year1.find('h3').text.strip()
    house_info['address'] = address1.find('h1').text.strip()
    house_info['area'] = area1.find('h3').text.strip()
    house_info['rooms'] = rooms1.find('h2').text.strip()
    print(house_info)

    return house_info


def send_telegram_message(message):
    api_token = "---telegram-Bot_API_token---"
    chat_id = "---telegram-chatID---"

    response = requests.post(f'https://api.telegram.org/bot{api_token}/sendMessage',
                             data={'chat_id': chat_id, 'text': message})

    if response.status_code == 200:
        print("Hyvä ilmoitus löydetty ja viesti lähetetty.")
    else:
        print(
            f"Hyvä ilmoitus löydetty, mutta viesti ei mennyt perille. ErrorCode: {response.status_code} - {response.content}")


def main():
    url = 'https://www.etuovi.com/myytavat-asunnot/kotka/kotkansaari?haku=M2088182874'
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')
    apartments_EndUrl = []

    house_urls = soup.find_all(class_="mui-style-1hvv1xy e3qdyeq2")
    for house in house_urls:
        if house is not None:
            houseURL = house["href"]
            apartments_EndUrl.append(houseURL)

    #print("x")
    print(*apartments_EndUrl, sep="\n")

    excel_file_path = "C:\\tests\\listings.xlsx"
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active

    for x in apartments_EndUrl:
        apartment_FullUrl = (f"https://www.etuovi.com{x}")
        print(apartment_FullUrl)
        house_info = get_house_info(apartment_FullUrl)

        if apartment_FullUrl in [cell.value for cell in sheet['B']]:
            print("Löytyy jo, skipataan seuraavaan")
            continue

        row = sheet.max_row + 1
        #sheet.cell(row=row, column=1, value=datetime)
        sheet.cell(row=row, column=2, value=apartment_FullUrl)
        sheet.cell(row=row, column=3, value=house_info['price'])
        sheet.cell(row=row, column=4, value=house_info['year_built'])
        sheet.cell(row=row, column=5, value=house_info['address'])
        sheet.cell(row=row, column=6, value=house_info['area'])
        sheet.cell(row=row, column=7, value=house_info['rooms'])

            #price = int(''.join(filter(str.isdigit, house_info['price'])))
            #area = int(re.search(r'\d+', house_info['area']).group())
            #year_built = int(''.join(filter(str.isdigit, house_info['year_built'])))
            #price_per_sqm = round(price / area)

            #sheet.cell(row=row, column=8, value=price_per_sqm)

        #if house_info['price'] <= 80000 and house_info['year_built'] >= 1950:
        message = (f"Uusi ilmoitus!\nHinta: {house_info['price']}\nRakennusvuosi: {house_info['year_built']}\n"
                   f"Neliöhinta: €\nURL: {apartment_FullUrl}")
        send_telegram_message(message)

    workbook.save(excel_file_path)
    workbook.close()


if __name__ == "__main__":
    main()